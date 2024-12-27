from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink

from data.output_file_styling_data import Colors


# TODO переделать на обычный класс, т.к. тут реализуется логика, а не хранятся данные
@dataclass
class ExcelStyling:
    output_file_path: str

    def __post_init__(self):
        self._wb = load_workbook(self.output_file_path)
        self._ws = self._wb.active

    @property
    def wb(self):
        return self._wb

    @property
    def ws(self):
        return self._ws

    def align_cells_horizontally(self, alignment: str = 'left'):
        alignment = Alignment(horizontal=alignment)
        for row in self.ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        self.wb.save(self.output_file_path)

    def set_column_width(self, width: str = None):
        self.ws.column_dimensions['A'].width = 50
        self.wb.save(self.output_file_path)

    def add_hyperlinks_to_indexes(self, indexes_list: list[str]):
        link = Hyperlink()
        pass

    def color_cells_by_test_results(self):
        green_fill = PatternFill(fill_type='solid', start_color=Colors.green, end_color=Colors.green)
        red_fill = PatternFill(fill_type='solid', start_color=Colors.red, end_color=Colors.red)

        for row in self.ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if cell.value == 'passed':
                    cell.fill = green_fill
                elif cell.value == 'failed' or cell.value == 'broken':
                    cell.fill = red_fill

        self.wb.save(self.output_file_path)
